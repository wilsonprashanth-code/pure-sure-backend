import openpyxl
import pandas as pd
import re

wb = openpyxl.load_workbook('base recipe.xlsx', data_only=True)

red_sheets = []
for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    color = sheet.sheet_properties.tabColor
    if color and color.rgb:
        color_str = str(color.rgb)
        if 'FF' in color_str and (color_str.endswith('0000') or color_str == 'FFFF0000'):
            red_sheets.append(sheet_name)

base_recipes_rows = []

for sheet_name in red_sheets:
    sheet = wb[sheet_name]
    data = [row for row in sheet.iter_rows(values_only=True)]
    df = pd.DataFrame(data)
    
    if df.empty or df.shape[0] < 3:
        continue
        
    recipe_name = str(df.iloc[1, 1]).strip() if df.shape[0] > 1 and df.iloc[1, 1] else sheet_name.strip()
    
    yield_info = str(df.iloc[2, 1]) if df.shape[0] > 2 and df.iloc[2, 1] else ""
    batch_yield = 1.0
    unit = "GRM"
    
    if "Batch Yield:" in yield_info:
        match = re.search(r'Batch Yield:\s*([\d\.]+)\s*([A-Za-z]+)', yield_info)
        if match:
            batch_yield = float(match.group(1))
            unit = match.group(2).upper()
            if unit in ['GM', 'G']: unit = 'GRM'
    
    header_idx = None
    for r in range(min(10, df.shape[0])):
        row_str = " ".join([str(val) for val in df.iloc[r, :].values if val is not None])
        if 'Ingredient Description' in row_str or 'Sl No.' in row_str:
            header_idx = r
            break
            
    if header_idx is not None:
        header_row = [str(cell).strip() if cell else "" for cell in df.iloc[header_idx, :]]
        
        ing_col, qty_col, unit_col, grams_col, uom_col, vendor_col = None, None, None, None, None, None
        
        for idx, col_name in enumerate(header_row):
            col_name_lower = col_name.lower()
            if 'ingredient description' in col_name_lower: ing_col = idx
            elif 'original qty' in col_name_lower or 'qty' in col_name_lower:
                if qty_col is None: qty_col = idx
            elif col_name_lower == 'unit': unit_col = idx
            elif 'grams' in col_name_lower or 'ml' in col_name_lower: grams_col = idx
            elif 'uom' in col_name_lower: uom_col = idx
            elif 'vendor' in col_name_lower: vendor_col = idx
                
        for r in range(header_idx + 1, df.shape[0]):
            ing_name = df.iloc[r, ing_col] if ing_col is not None and ing_col < df.shape[1] else None
            if not ing_name or str(ing_name).strip() == "" or "total" in str(ing_name).lower():
                continue
                
            qty_val = df.iloc[r, qty_col] if qty_col is not None and qty_col < df.shape[1] else 1
            unit_val = df.iloc[r, unit_col] if unit_col is not None and unit_col < df.shape[1] else "GRM"
            grams_val = df.iloc[r, grams_col] if grams_col is not None and grams_col < df.shape[1] else qty_val
            uom_val = df.iloc[r, uom_col] if uom_col is not None and uom_col < df.shape[1] else unit_val
            vendor_val = df.iloc[r, vendor_col] if vendor_col is not None and vendor_col < df.shape[1] else "FACTORY"
            
            try: qty_num = float(qty_val) if qty_val is not None else 1.0
            except: qty_num = 1.0
                
            try: grams_num = float(grams_val) if grams_val is not None else qty_num
            except: grams_num = qty_num
                
            base_recipes_rows.append({
                'recipeName': recipe_name,
                'batchYield': batch_yield,
                'unit': unit,
                'ingredientName': str(ing_name).strip(),
                'qty': qty_num,
                'ingredientUnit': str(unit_val).strip() if unit_val else "GRM",
                'gramsMl': grams_num,
                'uom': str(uom_val).strip() if uom_val else "GRM",
                'vendor': str(vendor_val).strip() if vendor_val else "FACTORY"
            })

df_export = pd.DataFrame(base_recipes_rows)
df_export.to_csv('Base_Recipes_49_Red_Sheets.csv', index=False)
print(f"Done! Base_Recipes_49_Red_Sheets.csv written with {len(df_export)} rows across {len(red_sheets)} red sheets.")
